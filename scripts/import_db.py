

# Load product information from excel sheet
# Excel sheet provided by Imran
import os,sys

from datetime import datetime, date, timedelta

import django
from django.core.files import File
from django.core.exceptions import ObjectDoesNotExist

from openpyxl import load_workbook

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
print(BASE_DIR)

if BASE_DIR not in sys.path:
	sys.path.append(BASE_DIR)

PRODUCT_CATEGORIES = (
	("servers", "Servers"),
	("storage", "Storage Appliances"),
	("network", "Networking"),
	("appliances", "Appliances"),
)

def get_cat(value, prods):
	for i in prods:
		if value==i.category_code:
			return i
	return False

SHEETS = [
	'DELL',
	'HP',
	'ORACLE',
	'CISCO',
	'EMC',
	'HITACHI',
	'NETAPP',
	'IBM',
	'SUPERMICRO',
]

def App2(csv):

# Main App
def App(wb, name):
	print("*************")
	print("***", name, "***")
	print("*************")

	product_categories = ProductCategory.objects.all()

	SHEET_NAME = name
	SHEET_LENGTH = wb[SHEET_NAME].max_row

	for row in range(1,SHEET_LENGTH):
		brand = wb[SHEET_NAME].cell(column=1, row=row).value
		model = wb[SHEET_NAME].cell(column=2, row=row).value
		category = wb[SHEET_NAME].cell(column=3, row=row).value
		print(row, ":", model, ":", category)
		cat_final = get_cat(category.lower(), product_categories.all())

		brand = brand.replace(" ", "")

		if model != None:
			# print(cat_final)
			# print(row, ": ", model)
			created = False
			try:
				prod = Product.objects.get(brand=brand, model=model)
			except ObjectDoesNotExist:
				created = True
				prod = Product(brand=brand, model=model, release=date.today() - timedelta(1), category=cat_final, price_silver=1.0, price_gold=1.0, price_black=1.0)

			if prod.category == None:
				prod.category = cat_final

			# DEV NOTES
			# Change SKU to iterate only within category
			# Not whole brand as it does now

			# Adjust SKU
			cat = ""
			if prod.category == "servers":
				cat = "SRV"
			elif prod.category == "storage":
				cat = "STG"
			elif prod.category == "network":
				cat = "NET"
			elif prod.category == "appliances":
				cat = "APP"
			prod.sku = prod.brand[0].upper() + cat + '-' + '{0:04d}'.format(row)
			
			prod.save()

			if created:
				print("-- Created --")
			else:
				print("** Loaded")
			print(brand, " ", model, " | ", prod.pk)
		else:
			print(row, " is empty")

# Django setup
if __name__ == '__main__':
	os.chdir(BASE_DIR)
	os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings.local')
	django.setup()
	from support.models import Product, ProductCategory

	for sheet in SHEETS:
		name = sheet.lower()
		workbook = load_workbook('scripts/files/products/mibura_sss_product_' + name + '.xlsx')
		App(workbook, sheet)